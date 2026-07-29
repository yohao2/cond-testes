import openpyxl
from dados import ibiza_agua, ibiza_gas, bali_agua, bali_gas
from dropbox_utils import salvar_no_dropbox


def gerar_planilhas(escolher_condominio):
    dict_condominios = {
        'Residencial Ibiza': gerar_planilhas_ibiza,
        'Residencial Bali': gerar_planilhas_ibiza
    }
    if escolher_condominio in dict_condominios:
        dict_condominios[escolher_condominio]()


def gerar_planilhas_ibiza(checkbox_agua, checkbox_gas, dia_mes_agua, mes_referencia, fatura, multa, lista_agua, nome_arquivo_agua, nome_arquivo_gas, dia_mes_gas, valor_bujao_unidade, valor_bujao_total, lista_gas ): 

    if checkbox_agua:

        wb = openpyxl.load_workbook(ibiza_agua)
        sheet = wb['AGUA']
        # dia do mes
        sheet.cell(row= 2, column = 5).value =  dia_mes_agua
        # mes de referencia
        sheet.cell(row=3, column = 5).value = mes_referencia

        # valor da fatura
        if multa:
            sheet.cell(row=27, column = 5).value = fatura - multa
            # valor da multa
            sheet.cell(row=27, column = 7).value = (f'{fatura} - {multa} DE MULTA')
        else: 
            sheet.cell(row=27, column = 5).value = fatura
            sheet.cell(row=27, column = 7).value = 'SEM MULTA'

        # leitura anterior 
        for linha in sheet.iter_rows(min_row=5, max_row = 25, values_only = False):
            linha[1].value = linha[2].value
    
        # mes de referencia
        sheet.cell(row=5,column=3).value = mes_referencia
        # data leitura
        sheet.cell(row = 6, column = 3).value = dia_mes_agua
        # leitura atual
        for i, valor in enumerate(lista_agua, start= 7):
            sheet.cell(row=i, column = 3).value = valor
    
        # calculo consumo
        for linha in sheet.iter_rows(min_row = 7, max_row = 25):
            linha[3].value = int(linha[2].value.replace(',','')) - int(linha[1].value.replace(',',''))
            linha[3].value = linha[3].value / 1000
    
        # verificar valor maior que 5m³
        for linha in sheet.iter_rows(min_row =7, max_row = 25,values_only = False):
            if int(linha[3].value) >= 5:
                linha[5].value = 'MAIOR QUE 5m³'
    
        # mudar nome da sheet
        wb['AGUA'].title = (mes_referencia)
        salvar_no_dropbox(wb, f'/I. {nome_arquivo_agua}.xlsx')


    if checkbox_gas:
        wb = openpyxl.load_workbook(ibiza_gas)
        sheet = wb['GAS']

        # dia do mes
        sheet.cell(row= 2, column = 6).value =  dia_mes_gas
        # mes de referencia
        sheet.cell(row=3, column = 6).value = mes_referencia
        # valor por bujão
        sheet.cell(row=7, column = 9).value = valor_bujao_unidade
        # valor da dos bujoes
        sheet.cell(row=27, column = 6).value =  valor_bujao_total
    
        # leitura anterior 
        for linha in sheet.iter_rows(min_row=5, max_row = 25, values_only = False):
            linha[1].value = linha[2].value
        # mes de referencia
        sheet.cell(row=5,column=3).value = mes_referencia
        # data leitura
        sheet.cell(row = 6, column = 3).value = dia_mes_gas
        # leitura atual
        for i, valor in enumerate(lista_gas, start= 7):
            sheet.cell(row=i, column = 3).value = valor
    
        # mudar nome da sheet
        wb['GAS'].title = (mes_referencia)
        salvar_no_dropbox(wb, f'/I. {nome_arquivo_gas}.xlsx')

def gerar_planilhas_bali(checkbox_agua, checkbox_gas, dia_mes_agua, mes_referencia, fatura, multa, lista_agua, nome_arquivo_agua, nome_arquivo_gas, dia_mes_gas, valor_bujao_unidade, valor_bujao_total, lista_gas):

    if checkbox_agua:
        wb = openpyxl.load_workbook(bali_agua)
        sheet = wb['AGUA']
        
        # dia do mes
        sheet.cell(row= 2, column = 5).value =  dia_mes_agua
        # mes de referencia
        sheet.cell(row=3, column = 5).value = mes_referencia
        # valor da fatura
        if multa:
            sheet.cell(row=28, column = 5).value = fatura - multa
            # valor da multa
            sheet.cell(row=28, column = 7).value = (f'{fatura} - {multa} DE MULTA')
            
        else: 
            sheet.cell(row=28, column = 5).value = fatura
            sheet.cell(row=28, column = 7).value = 'SEM MULTA'
    
    
        # leitura anterior 
        for linha in sheet.iter_rows(min_row=5, max_row = 26, values_only = False):
            linha[1].value = linha[2].value
    
        # mes de referencia
        sheet.cell(row=5,column=3).value = mes_referencia
    
        # data leitura
        sheet.cell(row = 6, column = 3).value = dia_mes_agua
    
        # leitura atual
        for i, valor in enumerate(lista_agua, start= 7):
            sheet.cell(row=i, column = 3).value = valor
    
        # calculo consumo
        for linha in sheet.iter_rows(min_row = 7, max_row = 26):
            linha[3].value = int(linha[2].value.replace(',','')) - int(linha[1].value.replace(',',''))
    
            linha[3].value = linha[3].value / 1000
    
        # verificar valor maior que 5m³
        for linha in sheet.iter_rows(min_row =7, max_row = 26,values_only = False):
            if int(linha[3].value) >= 5:
                linha[5].value = 'MAIOR QUE 5m³'
    
        salvar_no_dropbox(wb, f'/I. {nome_arquivo_agua}.xlsx')


    if checkbox_gas:
        wb = openpyxl.load_workbook(bali_agua)
        sheet = wb['GAS']

        # dia do mes
        sheet.cell(row= 2, column = 6).value =  dia_mes_gas
        # mes de referencia
        sheet.cell(row=3, column = 6).value = mes_referencia
        # valor por bujão
        sheet.cell(row=7, column = 9).value = valor_bujao_unidade
        # valor da dos bujoes
        sheet.cell(row=28, column = 6).value =  valor_bujao_total

        # leitura anterior 
        for linha in sheet.iter_rows(min_row=5, max_row = 26, values_only = False):
            linha[1].value = linha[2].value
        # mes de referencia
        sheet.cell(row=5,column=3).value = mes_referencia
        # data leitura
        sheet.cell(row = 6, column = 3).value = dia_mes_gas
        # leitura atual
        for i, valor in enumerate(lista_gas, start= 7):
            sheet.cell(row=i, column = 3).value = valor

        # mudar nome da sheet
        wb['GAS'].title = (mes_referencia)
        salvar_no_dropbox(wb, f'B. /{nome_arquivo_gas}.xlsx')

